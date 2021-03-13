import openpyxl
import os

from muitls import gen_months, Citys

excel_path = './excel/'


def rd_month_t4_t5(year, month, city, sheet=5, col="D"):
    assert city in Citys

    """
    表5：2015年1月70个大中城市二手住宅分类价格指数									
    城市	90m2及以下			90-144m2			144m2以上		
    环比	同比	定基	环比	同比	定基	环比	同比	定基
    上月=100	去年同月=100	2010年=100	上月=100	去年同月=100	2010年=100	上月=100	去年同月=100	2010年=100
    """
    filename = excel_path + \
        "t_{0}{1}.xlsx".format(year, '%2s' % (str(month).zfill(2)))
    if os.path.exists(filename):
        mexcel = openpyxl.load_workbook(filename)
        if year == 2017 or (year==2016 and month>8):
            assert len(mexcel._sheets) == 6
            sheet4_or_5 = mexcel._sheets[sheet]
            table_name = sheet4_or_5["A1"].internal_value
            assert "表"+str(sheet+1) in table_name and \
                str(year)+"年" in table_name and \
                str(month)+"月" in table_name
                # "二手" in table_name and "分类" in table_name  
        elif year < 2018 or (year==2018 and month < 3):
            assert len(mexcel._sheets) == 5
            sheet4_or_5 = mexcel._sheets[sheet-1]
            table_name = sheet4_or_5["A1"].internal_value
            assert "表"+str(sheet) in table_name and \
                str(year)+"年" in table_name and \
                str(month)+"月" in table_name
                # "二手" in table_name and "分类" in table_name
        else:
            assert len(mexcel._sheets) >= 4
            sheet4_or_5 = mexcel._sheets[sheet-2]
            table_name = sheet4_or_5["A1"].internal_value
            assert "表"+str(sheet-1) in table_name and \
                str(year)+"年" in table_name and \
                str(month)+"月" in table_name
                # "二手" in table_name and "分类" in table_name


        for row in sheet4_or_5.iter_rows():
            if isinstance(row[0], openpyxl.cell.cell.Cell) and row[0].row > 3 and row[0].internal_value:
            # if isinstance(row[0], openpyxl.cell.cell.Cell) and \
            #     isinstance(sheet4_or_5[col+str(row[0].row)], openpyxl.cell.cell.Cell) and \
            #         row[0].row > 3 and row[0].internal_value and sheet4_or_5[col+str(row[0].row)] and \
            #         sheet4_or_5[col+str(row[0].row)].internal_value:
                tcity = row[0].internal_value.strip(
                    " ").replace("　", '').replace(" ", "")
                # assert city in Citys
                if tcity == city:
                    d = float(sheet4_or_5[col+str(row[0].row)].internal_value)
                    if year <= 2015:
                        return d * (100/122.9)
                    else:
                        return d 

def judge_data(year, month,stype, sheet=5):

    """
    表5：2015年1月70个大中城市二手住宅分类价格指数									
    城市	90m2及以下			90-144m2			144m2以上		
    环比	同比	定基	环比	同比	定基	环比	同比	定基
    上月=100	去年同月=100	2010年=100	上月=100	去年同月=100	2010年=100	上月=100	去年同月=100	2010年=100
    """
    filename = excel_path + \
        "t_{0}{1}.xlsx".format(year, '%2s' % (str(month).zfill(2)))
    if os.path.exists(filename):
        mexcel = openpyxl.load_workbook(filename)
        if year == 2017 or (year==2016 and month>8):
            assert len(mexcel._sheets) == 6
            sheet4_or_5 = mexcel._sheets[sheet]
            table_name = sheet4_or_5["A1"].internal_value
            assert "表"+str(sheet+1) in table_name and \
                str(year)+"年" in table_name and \
                str(month)+"月" in table_name and \
                stype in table_name and "分类" in table_name  
        elif year < 2018 or (year==2018 and month < 3):
            assert len(mexcel._sheets) == 5
            sheet4_or_5 = mexcel._sheets[sheet-1]
            table_name = sheet4_or_5["A1"].internal_value
            assert "表"+str(sheet) in table_name and \
                str(year)+"年" in table_name and \
                str(month)+"月" in table_name and \
                stype in table_name and "分类" in table_name
        else:
            assert len(mexcel._sheets) >= 4
            sheet4_or_5 = mexcel._sheets[sheet-2]
            table_name = sheet4_or_5["A1"].internal_value
            assert "表"+str(sheet-1) in table_name and \
                str(year)+"年" in table_name and \
                str(month)+"月" in table_name and \
                stype in table_name and "分类" in table_name



def main():
    for month in gen_months('201601', '202012'):
        judge_data(month[0], month[1],'二手',sheet=5)
        print(month[0],month[1],"success")

if __name__ == "__main__":
    main()
